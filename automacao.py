import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
from datetime import datetime

# Configuração do Selenium
chrome_options = Options()
chrome_options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=chrome_options)

# URLs
SITE_URL = "https://sorrisobanguela-odonto.netlify.app"
DADOS_FILE = r"c:\Users\jonat\OneDrive\Área de Trabalho\projeto\Finanças_clinica_odonto\dados_clientes.xlsx"
FECHAMENTO_FILE = r"c:\Users\jonat\OneDrive\Área de Trabalho\projeto\Finanças_clinica_odonto\planilha fechamento.xlsx"

def carregar_dados_clientes():
    """Carrega os dados da planilha de clientes"""
    wb = openpyxl.load_workbook(DADOS_FILE)
    ws = wb.active
    
    clientes = []
    # Começa na linha 2 (linha 1 é cabeçalho)
    # Colunas: Nome | Valor | CPF | Forma de Pagamento | Data da Consulta | Status
    for row in ws.iter_rows(min_row=2, values_only=False):
        nome_cell = row[0]      # Nome
        valor_cell = row[1]     # Valor
        cpf_cell = row[2]       # CPF
        forma_cell = row[3]     # Forma de Pagamento
        data_cell = row[4]      # Data da Consulta
        status_cell = row[5]    # Status
        
        # Extrai valores
        if nome_cell.value and cpf_cell.value:
            cliente = {
                'nome': str(nome_cell.value).strip(),
                'valor': str(valor_cell.value).strip() if valor_cell.value else '',
                'cpf': str(cpf_cell.value).strip() if cpf_cell.value else '',
                'forma': str(forma_cell.value).strip() if forma_cell.value else '',
                'data_consulta': str(data_cell.value).strip() if data_cell.value else '',
                'status': str(status_cell.value).strip() if status_cell.value else ''
            }
            clientes.append(cliente)
    
    wb.close()
    return clientes

def formatar_cpf(cpf):
    """Formata CPF para o padrão XXX.XXX.XXX-XX"""
    cpf_digits = ''.join(filter(str.isdigit, cpf))
    if len(cpf_digits) == 11:
        return f"{cpf_digits[0:3]}.{cpf_digits[3:6]}.{cpf_digits[6:9]}-{cpf_digits[9:11]}"
    return cpf

def formatar_data_entrada(data_str):
    """Converte data para formato YYYY-MM-DD (entrada no HTML5 input date)"""
    if not data_str:
        return ''
    
    # Se já está em formato YYYY-MM-DD (correto)
    if '-' in data_str and len(data_str) >= 10:
        # Extrai apenas os primeiros 10 caracteres (YYYY-MM-DD)
        return data_str[:10]
    
    # Se está em formato DD/MM/YYYY
    if '/' in data_str:
        try:
            dt = datetime.strptime(data_str, '%d/%m/%Y')
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass
    
    # Trata datetime string genérica
    try:
        # Remove a parte de hora se existir (formato: 2026-04-07 00:00:00)
        if ' ' in data_str:
            data_str = data_str.split(' ')[0]
        
        # Se ainda tem hífen, já está no formato certo
        if '-' in data_str and len(data_str) == 10:
            return data_str
    except:
        pass
    
    return data_str

def formatar_data_saida(data_str):
    """Converte data para formato DD/MM/YYYY (saída para planilha)"""
    if not data_str:
        return ''
    
    try:
        # Se está em formato YYYY-MM-DD
        if '-' in data_str:
            data_str = data_str.split(' ')[0]  # Remove hora se tiver
            dt = datetime.strptime(data_str, '%Y-%m-%d')
            return dt.strftime('%d/%m/%Y')
        
        # Se está em formato DD/MM/YYYY já
        if '/' in data_str and len(data_str) == 10:
            return data_str
    except:
        pass
    
    return data_str

def formatar_valor(valor_str):
    """Limpa e retorna valor numérico sem formatação"""
    # Remove R$, espaços e converte vírgula em ponto
    valor_clean = valor_str.replace('R$', '').strip()
    valor_clean = valor_clean.replace('.', '').replace(',', '.')
    return valor_clean

def mapear_forma_pagamento(forma):
    """Mapeia forma de pagamento para o valor do select"""
    forma_lower = forma.lower().strip()
    if 'pix' in forma_lower:
        return 'pix'
    elif 'cartão' in forma_lower or 'cartao' in forma_lower:
        return 'cartao'
    elif 'boleto' in forma_lower:
        return 'boleto'
    elif 'dinheiro' in forma_lower or 'cash' in forma_lower:
        return 'dinheiro'
    return forma_lower

def navegar_para_cadastro():
    """Navega para a seção de cadastro"""
    try:
        botao_cadastro = driver.find_element(By.XPATH, "//button[contains(@onclick, \"navigate('cadastro')\")]")
        botao_cadastro.click()
        time.sleep(1)
    except:
        print("Erro ao navegar para cadastro")

def preencher_formulario(cliente):
    """Preenche o formulário com dados do cliente"""
    try:
        # Aguarda carregamento dos elementos
        wait = WebDriverWait(driver, 10)
        
        # Campo Nome
        campo_nome = wait.until(EC.presence_of_element_located((By.ID, "f-nome")))
        campo_nome.clear()
        campo_nome.send_keys(cliente['nome'])
        
        # Campo CPF
        campo_cpf = driver.find_element(By.ID, "f-cpf")
        campo_cpf.clear()
        cpf_formatado = formatar_cpf(cliente['cpf'])
        campo_cpf.send_keys(cpf_formatado)
        
        # Select Forma de Pagamento
        select_forma = Select(driver.find_element(By.ID, "f-forma"))
        forma_mapeada = mapear_forma_pagamento(cliente['forma'])
        select_forma.select_by_value(forma_mapeada)
        
        # Campo Valor
        campo_valor = driver.find_element(By.ID, "f-valor")
        campo_valor.clear()
        valor_limpo = formatar_valor(cliente['valor'])
        campo_valor.send_keys(valor_limpo)
        
        # Campo Data de Consulta (Vencimento) - Usar JavaScript pois input type="date" não aceita send_keys
        data_formatada = formatar_data_entrada(cliente['data_consulta'])
        campo_venc = driver.find_element(By.ID, "f-venc")
        driver.execute_script(f"document.getElementById('f-venc').value = '{data_formatada}'")
        
        # Marca a caixa de pago se o status do cliente for pago
        campo_pago = driver.find_element(By.ID, "f-pago")
        status_lower = cliente['status'].lower()
        if 'pago' in status_lower:
            if not campo_pago.is_selected():
                campo_pago.click()
        else:
            if campo_pago.is_selected():
                campo_pago.click()
        
        time.sleep(0.5)
        return True
    except Exception as e:
        print(f"Erro ao preencher formulário: {str(e)}")
        return False

def clicar_botao_salvar():
    """Clica no botão de salvar"""
    try:
        botao_salvar = driver.find_element(By.ID, "btn-salvar")
        botao_salvar.click()
        time.sleep(2)
        return True
    except Exception as e:
        print(f"Erro ao clicar no botão salvar: {str(e)}")
        return False

def coletar_informacoes_site():
    """Coleta as informações que foram cadastradas no site"""
    try:
        wait = WebDriverWait(driver, 5)
        
        # Tenta buscar as informações da última linha ou do formulário
        time.sleep(1)
        
        # Verifica se há mensagem de sucesso ou busca dados na tabela
        try:
            # Busca a última linha da tabela
            tabela = driver.find_element(By.TAG_NAME, "table")
            linhas = tabela.find_elements(By.TAG_NAME, "tr")
            ultima_linha = linhas[-1]
            
            colunas = ultima_linha.find_elements(By.TAG_NAME, "td")
            
            if len(colunas) >= 6:
                # Extrai informações esperadas
                info = {
                    'vencimento': formatar_data_saida(colunas[4].text) if len(colunas) > 4 else 'N/A',
                    'status': colunas[5].text if len(colunas) > 5 else 'N/A',
                    'data_pagamento': datetime.now().strftime('%d/%m/%Y'),
                    'metodo_pagamento': colunas[3].text if len(colunas) > 3 else 'N/A'
                }
                return info
        except:
            pass
        
        # Se não conseguir da tabela, retorna dados do formulário + data atual
        info = {
            'vencimento': formatar_data_saida(driver.find_element(By.ID, "f-venc").get_attribute("value")) or 'N/A',
            'status': 'Cadastrado',
            'data_pagamento': datetime.now().strftime('%d/%m/%Y'),
            'metodo_pagamento': driver.find_element(By.ID, "f-forma").get_attribute("value") or 'N/A'
        }
        return info
    except Exception as e:
        print(f"Erro ao coletar informações: {str(e)}")
        return {
            'vencimento': 'N/A',
            'status': 'Erro ao coletar',
            'data_pagamento': datetime.now().strftime('%d/%m/%Y'),
            'metodo_pagamento': 'N/A'
        }

def salvar_dados_fechamento(cliente, info_site):
    """Salva os dados na planilha de fechamento"""
    try:
        # Carrega ou cria a planilha de fechamento
        try:
            wb = openpyxl.load_workbook(FECHAMENTO_FILE)
            ws = wb.active
            # Encontra a próxima linha vazia
            ultima_linha = ws.max_row + 1
        except:
            # Se não existir, cria nova
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Fechamento"
            ultima_linha = 1
            
            # Adiciona cabeçalhos na ordem correta
            # Colunas: Nome | Valor | CPF | Vencimento | Status | Data pagamento | Método pagamento
            cabecalhos = ['Nome', 'Valor', 'CPF', 'Vencimento', 'Status', 'Data pagamento', 'Método pagamento']
            for col, cabecalho in enumerate(cabecalhos, 1):
                ws.cell(row=1, column=col, value=cabecalho)
            ultima_linha = 2
        
        # Adiciona dados do cliente na ordem correta
        ws.cell(row=ultima_linha, column=1, value=cliente['nome'])
        ws.cell(row=ultima_linha, column=2, value=cliente['valor'])
        ws.cell(row=ultima_linha, column=3, value=cliente['cpf'])
        ws.cell(row=ultima_linha, column=4, value=info_site['vencimento'])
        ws.cell(row=ultima_linha, column=5, value=info_site['status'])
        ws.cell(row=ultima_linha, column=6, value=info_site['data_pagamento'])
        ws.cell(row=ultima_linha, column=7, value=info_site['metodo_pagamento'])
        
        # Salva o arquivo
        wb.save(FECHAMENTO_FILE)
        wb.close()
        print(f"✓ Dados de {cliente['nome']} salvos em fechamento")
        return True
    except Exception as e:
        print(f"Erro ao salvar dados de fechamento: {str(e)}")
        return False

def main():
    """Função principal"""
    try:
        print("Iniciando automação...")
        
        # Abre o site
        driver.get(SITE_URL)
        print(f"Abrindo {SITE_URL}")
        time.sleep(2)
        
        # Carrega dados dos clientes
        clientes = carregar_dados_clientes()
        print(f"Carregados {len(clientes)} clientes")
        
        # Processa cada cliente
        for idx, cliente in enumerate(clientes, 1):
            print(f"\n--- Processando cliente {idx}/{len(clientes)}: {cliente['nome']} ---")
            
            # Navega para cadastro
            navegar_para_cadastro()
            
            # Preenche formulário
            if preencher_formulario(cliente):
                print(f"✓ Formulário preenchido para {cliente['nome']}")
                
                # Clica em salvar
                if clicar_botao_salvar():
                    print(f"✓ Cadastro salvo para {cliente['nome']}")
                    
                    # Coleta informações do site
                    info_site = coletar_informacoes_site()
                    print(f"✓ Informações coletadas: Status={info_site['status']}")
                    
                    # Salva na planilha de fechamento
                    salvar_dados_fechamento(cliente, info_site)
                else:
                    print(f"✗ Erro ao salvar cadastro de {cliente['nome']}")
            else:
                print(f"✗ Erro ao preencher formulário para {cliente['nome']}")
            
            time.sleep(1)
        
        print("\n✓ Automação concluída com sucesso!")
        input("Pressione Enter para encerrar...")
        
    except Exception as e:
        print(f"Erro geral: {str(e)}")
        input("Pressione Enter para encerrar...")
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
