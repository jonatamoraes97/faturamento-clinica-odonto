import time
import os
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from supabase import create_client, Client

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# ─── CREDENCIAIS (.env) ───────────────────────────────────────────────────────
load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise EnvironmentError("Defina SUPABASE_URL e SUPABASE_KEY no arquivo .env")

sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ─── CONFIGURAÇÕES ────────────────────────────────────────────────────────────
URL_SITE           = "https://sorrisobanguela-odonto.netlify.app"
BASE_DIR           = Path(__file__).parent
ARQUIVO_CLIENTES   = BASE_DIR / "dados_clientes.xlsx"
ARQUIVO_FECHAMENTO = BASE_DIR / "planilha_fechamento.xlsx"
TIMEOUT            = 15

# ─── ÍNDICES DE COLUNA — dados_clientes.xlsx
# Nome | Valor | CPF | Forma de Pagamento | Data da Consulta | Status
COL_NOME   = 0
COL_VALOR  = 1
COL_CPF    = 2
COL_FORMA  = 3
COL_DATA   = 4
COL_STATUS = 5

# ─── CABEÇALHO EXATO — planilha_fechamento.xlsx
# Nome | Valor | CPF | Vencimento | Status | Data pagamento | Método pagamento
CABECALHO_FECHAMENTO = [
    "Nome", "Valor", "CPF",
    "Vencimento", "Status",
    "Data pagamento", "Método pagamento",
]

# ─── MAPEAMENTO forma de pagamento → valor do <option> no site
FORMA_MAPA = {
    "pix":                  "pix",
    "cartão de crédito":    "cartao_credito",
    "cartao de credito":    "cartao_credito",
    "cartão crédito":       "cartao_credito",
    "cartao credito":       "cartao_credito",
    "crédito":              "cartao_credito",
    "credito":              "cartao_credito",
    "cartão de débito":     "cartao_debito",
    "cartao de debito":     "cartao_debito",
    "cartão débito":        "cartao_debito",
    "cartao debito":        "cartao_debito",
    "débito":               "cartao_debito",
    "debito":               "cartao_debito",
    "boleto":               "boleto",
    "dinheiro":             "dinheiro",
    "espécie":              "dinheiro",
    "especie":              "dinheiro",
}

# ─── CORES ────────────────────────────────────────────────────────────────────
COR_CABECALHO = "1E3A5F"
COR_PAGO      = "DCFCE7"
COR_ATRASADO  = "FEE2E2"
COR_PENDENTE  = "FEF3C7"
COR_ERRO      = "F3F4F6"

def _borda():
    s = Side(border_style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── CONVERSÃO DE DATA ───────────────────────────────────────────────────────

def converter_data(valor) -> str:
    """
    Converte para YYYY-MM-DD qualquer formato que o openpyxl devolva:
    - datetime / date já convertido pelo openpyxl
    - número serial do Excel (ex: 46119 → 2026-04-07)
    - string dd/mm/yyyy, yyyy-mm-dd, etc.
    """
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(valor))).strftime("%Y-%m-%d")
        except (ValueError, OverflowError):
            return ""
    texto = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return texto


def normalizar(texto: str) -> str:
    return (str(texto).strip().lower()
            .replace("á","a").replace("ã","a").replace("â","a")
            .replace("é","e").replace("ê","e").replace("í","i")
            .replace("ó","o").replace("ô","o").replace("ú","u")
            .replace("ç","c"))


def mapear_forma(forma_raw: str) -> str:
    chave = normalizar(forma_raw)
    if chave in FORMA_MAPA:
        return FORMA_MAPA[chave]
    for k, v in FORMA_MAPA.items():
        if k in chave or chave in k:
            return v
    print(f"  [AVISO] Forma '{forma_raw}' não reconhecida — usando 'pix'")
    return "pix"

# ─── LEITURA DE dados_clientes.xlsx ──────────────────────────────────────────

def ler_clientes() -> list:
    if not ARQUIVO_CLIENTES.exists():
        raise FileNotFoundError(f"Não encontrado: {ARQUIVO_CLIENTES}")

    wb = openpyxl.load_workbook(ARQUIVO_CLIENTES, data_only=True)
    ws = wb.active
    clientes = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        nome       = str(row[COL_NOME]   or "").strip()
        cpf        = str(row[COL_CPF]    or "").strip()
        forma_raw  = str(row[COL_FORMA]  or "").strip()
        status_raw = str(row[COL_STATUS] or "").strip()

        if not nome or not cpf:
            continue

        v = row[COL_VALOR]
        valor = float(v) if isinstance(v, (int, float)) else float(
            str(v or "0").replace("R$","").replace(".","").replace(",",".").strip() or "0"
        )

        clientes.append({
            "nome":           nome,
            "cpf":            cpf,
            "valor":          valor,
            "forma":          mapear_forma(forma_raw),
            "forma_original": forma_raw,
            "data_consulta":  converter_data(row[COL_DATA]),
            "status":         status_raw,
        })

    wb.close()
    return clientes

# ─── SUPABASE ────────────────────────────────────────────────────────────────

def gravar_supabase(cliente: dict, dados_site: dict):
    """Insere o registro na tabela 'pagamentos' do Supabase."""
    # Converte vencimento "dd/mm/yyyy" → "yyyy-mm-dd" para o banco
    venc_raw = dados_site.get("vencimento", "")
    venc_db  = None
    if venc_raw and venc_raw not in ("N/A", "ERRO", ""):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                venc_db = datetime.strptime(venc_raw, fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

    registro = {
        "nome":            cliente["nome"],
        "cpf":             cliente["cpf"],
        "valor":           cliente["valor"],
        "forma_pagamento": cliente["forma_original"],
        "data_consulta":   cliente["data_consulta"] or None,
        "vencimento":      venc_db,
        "status":          dados_site.get("status", ""),
        "metodo_site":     dados_site.get("metodo", ""),
        "status_planilha": cliente["status"],
    }

    try:
        sb.table("pagamentos").insert(registro).execute()
        print(f"          ✓ Gravado no Supabase")
    except Exception as e:
        print(f"          [SUPABASE ERRO] {e}")

# ─── SELENIUM ────────────────────────────────────────────────────────────────

def criar_driver(headless: bool = False) -> webdriver.Chrome:
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1440,900")
    opts.add_argument("--lang=pt-BR")
    driver = webdriver.Chrome(options=opts)
    driver.implicitly_wait(3)
    return driver


def aguardar(driver, by, seletor, clicavel=False, timeout=TIMEOUT):
    cond = EC.element_to_be_clickable if clicavel else EC.presence_of_element_located
    return WebDriverWait(driver, timeout).until(cond((by, seletor)))


def navegar_para_cadastro(driver):
    btn = aguardar(driver, By.XPATH,
                   "//button[contains(@onclick,\"navigate('cadastro')\")]",
                   clicavel=True)
    driver.execute_script("arguments[0].click();", btn)
    aguardar(driver, By.ID, "f-nome")
    time.sleep(0.4)


def cadastrar_cliente(driver, cliente: dict) -> dict:
    navegar_para_cadastro(driver)

    # Nome
    el = driver.find_element(By.ID, "f-nome")
    el.clear()
    el.send_keys(cliente["nome"])

    # CPF — dígitos puros; a máscara do site formata automaticamente
    el = driver.find_element(By.ID, "f-cpf")
    el.clear()
    el.send_keys(cliente["cpf"].replace(".", "").replace("-", "").strip())

    # Forma de pagamento
    Select(aguardar(driver, By.ID, "f-forma", clicavel=True)).select_by_value(cliente["forma"])
    time.sleep(0.3)

    # Valor — envia centavos para acionar maskValor corretamente
    centavos = str(int(round(cliente["valor"] * 100)))
    el = driver.find_element(By.ID, "f-valor")
    el.clear()
    el.send_keys(centavos)
    driver.execute_script(
        "document.getElementById('f-valor')"
        ".dispatchEvent(new Event('input',{bubbles:true}));"
    )

    # Data de consulta — via JS para evitar problemas com locale em type=date
    driver.execute_script(
        "var el=document.getElementById('f-venc');"
        "el.value=arguments[0];"
        "el.dispatchEvent(new Event('change',{bubbles:true}));"
        "el.dispatchEvent(new Event('input',{bubbles:true}));",
        cliente["data_consulta"]
    )
    time.sleep(0.3)

    # Clica em "Cadastrar Paciente"
    driver.execute_script(
        "arguments[0].click();",
        aguardar(driver, By.ID, "btn-salvar", clicavel=True)
    )
    time.sleep(1.4)

    return ler_linha_tabela(driver, cliente["cpf"])


def ler_linha_tabela(driver, cpf: str) -> dict:
    """
    Tabela do dashboard:
    Paciente(0) | CPF(1) | Valor(2) | Pagamento(3) | Vencimento(4) | Status(5) | Ações(6)
    """
    cpf_digits = cpf.replace(".", "").replace("-", "").strip()
    fallback   = {"vencimento": "N/A", "status": "Erro ao coletar", "metodo": "N/A"}

    try:
        WebDriverWait(driver, TIMEOUT).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "#dashboard-table table tbody tr")
            )
        )
    except TimeoutException:
        return fallback

    for linha in reversed(driver.find_elements(
            By.CSS_SELECTOR, "#dashboard-table table tbody tr")):
        cels = linha.find_elements(By.TAG_NAME, "td")
        if len(cels) < 6:
            continue
        if cels[1].text.replace(".", "").replace("-", "").strip() == cpf_digits:
            return {
                "metodo":     cels[3].text.strip(),
                "vencimento": cels[4].text.strip(),
                "status":     cels[5].text.strip(),
            }

    return fallback

# ─── PLANILHA DE FECHAMENTO ───────────────────────────────────────────────────

def preparar_fechamento():
    if ARQUIVO_FECHAMENTO.exists():
        wb = openpyxl.load_workbook(ARQUIVO_FECHAMENTO)
        ws = wb.active
        cab_atual = [str(c.value or "").strip()
                     for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if cab_atual == CABECALHO_FECHAMENTO:
            return wb, ws
        wb.remove(ws)
        ws = wb.create_sheet("Sheet1", 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    ws.append(CABECALHO_FECHAMENTO)
    for col in range(1, len(CABECALHO_FECHAMENTO) + 1):
        cel = ws.cell(1, col)
        cel.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cel.fill      = PatternFill("solid", fgColor=COR_CABECALHO)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border    = _borda()
    ws.row_dimensions[1].height = 28

    for i, w in enumerate([28, 12, 18, 16, 22, 18, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb, ws


def cor_linha(status: str) -> str:
    s = status.lower()
    if "pago" in s and "não" not in s and "nao" not in s:
        return COR_PAGO
    if "atraso" in s or "atrasado" in s:
        return COR_ATRASADO
    if "erro" in s or "n/a" in s:
        return COR_ERRO
    return COR_PENDENTE


def escrever_linha(ws, linha: int, cliente: dict, dados_site: dict):
    status_site = dados_site.get("status", "")
    cor = cor_linha(status_site)

    valores = [
        cliente["nome"],
        cliente["valor"],
        cliente["cpf"],
        dados_site.get("vencimento", ""),
        status_site,
        datetime.now().strftime("%d/%m/%Y"),
        dados_site.get("metodo", ""),
    ]

    for col, val in enumerate(valores, 1):
        cel = ws.cell(linha, col, value=val)
        cel.font      = Font(name="Arial", size=10)
        cel.alignment = Alignment(
            horizontal="left" if col == 1 else "center",
            vertical="center"
        )
        cel.fill   = PatternFill("solid", fgColor=cor)
        cel.border = _borda()

    ws.cell(linha, 2).number_format = 'R$ #,##0.00'
    ws.row_dimensions[linha].height = 18

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  OdontoGestão — Automação de Cadastro de Pagamentos")
    print("=" * 60)

    print(f"\n[1/4] Lendo clientes: {ARQUIVO_CLIENTES.name}")
    clientes = ler_clientes()
    print(f"      {len(clientes)} cliente(s) encontrado(s)\n")
    for c in clientes:
        print(f"      • {c['nome']:<35} {c['cpf']}  "
              f"R${c['valor']:>8.2f}  {c['forma_original']:<20}  {c['data_consulta']}")

    if not clientes:
        print("[ERRO] Nenhum cliente encontrado. Encerrando.")
        return

    print(f"\n[2/4] Preparando planilha: {ARQUIVO_FECHAMENTO.name}")
    wb_fech, ws_fech = preparar_fechamento()
    proxima_linha = ws_fech.max_row + 1
    print(f"      Próxima linha de escrita: {proxima_linha}")

    print(f"\n[3/4] Conectado ao Supabase: {SUPABASE_URL}")

    print(f"\n[4/4] Abrindo navegador → {URL_SITE}\n")
    driver = criar_driver(headless=False)   # headless=True para rodar sem janela

    try:
        driver.get(URL_SITE)
        aguardar(driver, By.CSS_SELECTOR, ".sidebar-logo .brand")
        print(f"      Site carregado: {driver.title}\n")
        time.sleep(1)

        total = len(clientes)
        for i, cliente in enumerate(clientes, 1):
            print(f"  [{i:>3}/{total}] {cliente['nome']}")
            try:
                dados_site = cadastrar_cliente(driver, cliente)
                print(f"          Vencimento : {dados_site['vencimento']}")
                print(f"          Status     : {dados_site['status']}")
                print(f"          Método     : {dados_site['metodo']}")
            except Exception as e:
                print(f"          [ERRO SELENIUM] {e}")
                dados_site = {"vencimento": "ERRO", "status": str(e)[:60], "metodo": ""}
                try:
                    driver.get(URL_SITE)
                    aguardar(driver, By.CSS_SELECTOR, ".sidebar-logo .brand")
                    time.sleep(1)
                except Exception:
                    pass

            # Grava no Supabase
            gravar_supabase(cliente, dados_site)

            # Grava na planilha Excel
            escrever_linha(ws_fech, proxima_linha, cliente, dados_site)
            proxima_linha += 1
            wb_fech.save(ARQUIVO_FECHAMENTO)

    finally:
        driver.quit()

    print("\n" + "=" * 60)
    print(f"  Concluído! {proxima_linha - 2} registro(s) gravado(s).")
    print(f"  → Excel:    {ARQUIVO_FECHAMENTO}")
    print(f"  → Supabase: {SUPABASE_URL}/project/default/editor")
    print("=" * 60)


if __name__ == "__main__":
    main()
